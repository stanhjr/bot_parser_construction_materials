from openpyxl.reader import excel
from openpyxl.styles import PatternFill
from openpyxl.styles import Font


def is_float_or_zero(price: str):
    try:
        price = float(price)
        if price == 0:
            return ''
        return price
    except ValueError:
        return ''


def set_price_to_book(address, price, file_path):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    try:
        sheet[address] = is_float_or_zero(price)
        wb.save(file_path)
        return True
    finally:
        wb.save(file_path)


async def get_link_epicenter(last_row: int, file_path):

    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    counter = 2
    liter = 'L'
    list_all_market = []

    while counter != last_row:
        if sheet[liter + str(counter)].value:
            one_market = sheet[liter + str(counter)].value, sheet[liter + str(counter)].coordinate,
            list_all_market.append(one_market)
        counter += 1
    return list_all_market


def get_all_link_market(last_row: int, file_path):

    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    counter = 2
    row_excel_book = 'DEFGHIJKLMNOPQR'
    list_all_market = []

    while counter != last_row:
        for liter in row_excel_book:
            one_market = sheet[liter + str(counter)].value, sheet[liter + str(counter)].coordinate, sheet[liter + str(counter)].coordinate[0]
            list_all_market.append(one_market)
        counter += 1
    return list_all_market


async def find_last_row_excel(file_path):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    row_excel_book = 'DEFGHIJKLMNOPQR'
    row = 2
    while True:
        row += 1
        for cell_number in range(row, row + 1):
            counter = 0
            for liter in row_excel_book:
                if sheet[liter + str(cell_number)].value is None:
                    counter += 1
                if counter == len(row_excel_book):
                    return int(sheet[liter + str(cell_number)].coordinate[1:]) * len(row_excel_book), int(sheet[liter + str(cell_number)].coordinate[1:])


def is_float(string: str):
    try:
        if isinstance(string, str):
            string = string.replace(',', '.')
        result = float(string)
        return result
    except Exception as e:
        return ''


def cut_paste_cell(file_path, last_row):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    list_cell_number = [x for x in range(1, last_row + 1)]
    list_cell_number.reverse()
    row_excel_book = 'RQPONMLKJIHG'
    for num in list_cell_number:
        for idx in range(len(row_excel_book)):
            if idx > 0:
                cell_old = row_excel_book[idx] + str(num)
                cell_new = row_excel_book[idx-1] + str(num)
                sheet[cell_new] = sheet[cell_old].value

    sheet["G1"] = "dnipro.atlant-shop.com.ua opt"
    wb.save(file_path)


def cut_paste_atlant_shop(file_path, last_row):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active
    list_cell_number = [x for x in range(2, last_row + 1)]
    for num in list_cell_number:
        g_cell = 'G' + str(num)
        r_cell = 'R' + str(num)
        sheet[g_cell] = sheet[r_cell].value
        sheet[r_cell] = ''
    sheet['Q1'].font = Font(bold=True)
    wb.save(file_path)


def markup_excel_file(file_path, last_row):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active

    try:
        for i in range(2, last_row + 1):
            row_excel_book = 'BEFGHIJKLMNOPRQ'
            d_cell = 'D' + str(i)
            d_value = is_float(sheet[d_cell].value)

            if not d_value:
                continue

            for letter in row_excel_book:
                cell = letter + str(i)
                cell_value = is_float(sheet[cell].value)
                if letter == 'E' and not cell_value:
                    sheet[cell] = ''
                elif letter == 'B':
                    sheet[cell] = ''
                elif letter == 'G' and not cell_value:
                    sheet[cell] = ''
                if not cell_value:
                    continue
                elif cell_value > d_value:
                    sheet[cell].fill = PatternFill(start_color="BEE8C4", fill_type="solid")
                elif cell_value < d_value:
                    sheet[cell].fill = PatternFill(start_color="F7C5C5", fill_type="solid")

        wb.save(file_path)
        return file_path
    finally:
        wb.save(file_path)


def markup_g_sell(file_path, last_row):
    wb = excel.load_workbook(file_path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    wb.active = 0
    sheet = wb.active

    try:
        for i in range(2, last_row + 1):
            cell = 'G' + str(i)
            value = is_float(sheet[cell].value)
            sheet[cell] = value
        # wb.save(file_path)
    finally:
        wb.save(file_path)

